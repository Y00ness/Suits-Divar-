from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import openpyxl, os, time


#################################
############# defs ##############
#################################

# def for input number of suit you want to see
def Number_of_suits():
    while True:
        Number_of_suit = input('How many suits you wanna see? (less than 150):')
        try:
            if int(Number_of_suit) in range(1, 151):
                return int(Number_of_suit)
            else:
                print('Select a Nume from 0 to 150: ')
        except:
            print('Select a Nume from 0 to 150: ')


# def for findin meter, rooms, price, owner and description from a suit page
def suit_info(res):
    soup = BeautifulSoup(res, 'lxml')
    post_info_block = soup.find('div', class_='post-info')
    meter_room_list = post_info_block.find_all('span', class_='kt-group-row-item__value')
    price_owner_list = post_info_block.find_all('div', class_='kt-base-row__end kt-unexpandable-row__value-box')
    description = post_info_block.find('p',
                                       class_='kt-description-row__text post-description kt-description-row__text--primary')
    return meter_room_list[0].text, meter_room_list[1].text, price_owner_list[0].text, price_owner_list[
        1].text, description.text


# def for select city
def city_url():
    while True:
        Input_num = input('Select Ur city --> ( 0:Kerman  1:Shiraz  2:Esfehan  3:Mashhad ) <--: ')
        if Input_num not in str([0, 1, 2, 3]):
            print('Wrong City...Try Again :)')
        else:
            print('Please wait...', end=' ')
            return int(Input_num)


# def for scrape main page
def scrape(i):
    all_blocks = soup.find('div', class_='browse-post-list')
    for block in all_blocks:
        try:
            link = block.find('a').attrs['href']
        except:
            pass
        if link not in list_of_links:
            list_of_links.append(link)
            suit_name = block.find('div', class_='kt-post-card__title').text
            try:
                time = block.find('span', class_='kt-post-card__bottom-description kt-text-truncate').text
            except:
                time = "فوری"
            # response suit page
            suit_sourcePage = f'https://divar.ir{link}'
            print(f'response page {i}...', end=' ')
            driver.get(suit_sourcePage)
            suit_res = driver.page_source
            try:
                meter = suit_info(suit_res)[0]
            except:
                meter = 'Not Available'
            try:
                rooms = suit_info(suit_res)[1]
            except:
                rooms = 'Not Available'
            try:
                price = suit_info(suit_res)[2]
            except:
                price = 'Not Available'
            try:
                owner = suit_info(suit_res)[3]
            except:
                owner = 'Not Available'
            try:
                description = suit_info(suit_res)[4]
            except:
                description = 'Not Available'
            # append info to the excel worksheet
            ws.append([suit_name, price, meter, rooms, time, owner, 'https://link', description])
            i += 1
            ws[f'G{i}'].hyperlink = suit_sourcePage
            print('Done')
        if (i - 1) == suit_numbers:
            return i - 1
    return i


# def for adjust cell's alignment  in excel_worksheet
def alignment_cell(ws):
    cell_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment.readingOrder = 2 #RTL
    for column in ws.columns:
        len_of_cell = 1
        for cell in column:
            cell.alignment = cell_alignment
            if cell.value:
                if len(cell.value) > len_of_cell:
                    len_of_cell = len(cell.value)
            ws.column_dimensions[get_column_letter(cell.column)].width = len_of_cell

#################################
############# main ##############
#################################

# define a driver
op = webdriver.ChromeOptions()
op.add_argument("window-size=1500,800")
op.add_argument('headless')
driver = webdriver.Chrome(executable_path='F:\ChromeDriver\chromedriver.exe', options=op)

# creat excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Divar Suits(Kerman)'
ws.sheet_view.rightToLeft = True
ws.append(['x سویت x', 'x قیمت x', 'x متراژ x', 'x اتاق x', 'x زمان آگهی x', 'x آگهی دهنده x', 'x لینک x', 'x توضیحات x'])

# list of city's urls
Kerman_url = 'https://divar.ir/s/kerman/rent-temporary-suite-apartment'
Esfehan_url = 'https://divar.ir/s/isfahan/rent-temporary-suite-apartment'
Mashhad_url = 'https://divar.ir/s/mashhad/rent-temporary-suite-apartment'
Shiraz_url = 'https://divar.ir/s/shiraz/rent-temporary-suite-apartment'
Urls_list = [Kerman_url, Shiraz_url, Esfehan_url, Mashhad_url]

# response main_page of Suits ####################
print('response main page ...', end=' ')
Selected_Url = Urls_list[city_url()]
try:
    driver.get(Selected_Url)
    print('Done')
except:
    print("Can't response..check your internet or other things... :(")

# select how many suits you want
suit_numbers = Number_of_suits()

# scrape suits
list_of_links = []
scraped_pages = 1
key_down = 1
while True:
    driver.get(Selected_Url)
    body = driver.find_element_by_tag_name('body')
    body.send_keys(Keys.CONTROL + Keys.HOME)
    time.sleep(1)
    for i in range(1, key_down):
        body.send_keys(Keys.PAGE_DOWN)
        time.sleep(2)
    key_down += 1
    res = driver.page_source
    soup = BeautifulSoup(res, 'lxml')
    scraped_pages = scrape(scraped_pages)
    print('please wait...')
    if scraped_pages == suit_numbers:
        break

# adjust worksheet ####################
alignment_cell(ws)
for cell in ws['H']:
    cell.alignment = Alignment(horizontal='right')
i = 1
for row in ws.rows:
    ws.row_dimensions[i].height = 20
    i += 1
for cell in ws[1]:
    cell.font = Font(bold=True, color='5F05AA')

# save and open excel file,close the driver #####
wb.save('Divar_Suits.xlsx')
os.system('start excel.exe Divar_suits.xlsx')
driver.close()

print('done')
i = input('\nHave a good day... :)\n\ndeveloped by ==Unes==')